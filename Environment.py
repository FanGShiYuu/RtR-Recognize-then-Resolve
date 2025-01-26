import math
import os
import random
import copy
import numpy as np
import openpyxl
from matplotlib import pyplot as plt
from scipy.interpolate import splrep, splev
from time import gmtime, strftime
from params import *

def RGB_to_Hex(rgb):
    RGB = rgb.split(',') 
    color = '#'
    for i in RGB:
        num = int(i)
        color += str(hex(num))[-2:].replace('x', '0').upper()
    print(color)
    return color

def scenario_outfit(ax, color=RGB_to_Hex('202,202,202')):
    # ax.plot(np.arange(15, 100, 1), 15 * np.ones(shape=(85,)), c=color)  
    # ax.plot(np.arange(15, 100, 1), -15 * np.ones(shape=(85,)), c=color)
    # ax.plot(np.arange(-99, -14, 1), 15 * np.ones(shape=(85,)), c=color)
    # ax.plot(np.arange(-99, -14, 1), -15 * np.ones(shape=(85,)), c=color)
    #
    # ax.plot(15 * np.ones(shape=(85,)), np.arange(15, 100, 1), c=color) 
    # ax.plot(-15 * np.ones(shape=(85,)), np.arange(15, 100, 1), c=color)
    # ax.plot(15 * np.ones(shape=(85,)), np.arange(-99, -14, 1), c=color)
    # ax.plot(-15 * np.ones(shape=(85,)), np.arange(-99, -14, 1), c=color)

    ax.plot(np.arange(15, 100, 1), 10 * np.ones(shape=(85,)), c=color)  
    ax.plot(np.arange(15, 100, 1), -10 * np.ones(shape=(85,)), c=color)
    ax.plot(np.arange(-99, -14, 1), 10 * np.ones(shape=(85,)), c=color)
    ax.plot(np.arange(-99, -14, 1), -10 * np.ones(shape=(85,)), c=color)

    ax.plot(10 * np.ones(shape=(85,)), np.arange(15, 100, 1), c=color)  
    ax.plot(-10 * np.ones(shape=(85,)), np.arange(15, 100, 1), c=color)
    ax.plot(10 * np.ones(shape=(85,)), np.arange(-99, -14, 1), c=color)
    ax.plot(-10 * np.ones(shape=(85,)), np.arange(-99, -14, 1), c=color)

    ax.plot(0 * np.ones(shape=(140,)), np.arange(-70, 70, 1), c='black', linestyle='--', linewidth=0.5,alpha=0.5) 
    ax.plot(np.arange(-70, 70, 1), 0 * np.ones(shape=(140,)), c='black', linestyle='--', linewidth=0.5, alpha=0.5)

def smooth_ployline(cv_init, point_num=3000):
    cv = cv_init
    list_x = cv[:, 0]
    list_y = cv[:, 1]
    if type(cv) is not np.ndarray:
        cv = np.array(cv)
    delta_cv = cv[1:, ] - cv[:-1, ]
    s_cv = np.linalg.norm(delta_cv, axis=1)

    s_cv = np.array([0] + list(s_cv))
    s_cv = np.cumsum(s_cv)
    bspl_x = splrep(s_cv, list_x, s=0.1)
    bspl_y = splrep(s_cv, list_y, s=0.1)
    # values for the x axis
    s_smooth = np.linspace(0, max(s_cv), point_num)
    # get y values from interpolated curve
    x_smooth = splev(s_smooth, bspl_x)
    y_smooth = splev(s_smooth, bspl_y)
    new_cv = np.array([x_smooth, y_smooth]).T
    delta_new_cv = new_cv[1:, ] - new_cv[:-1, ]
    s_accumulated = np.cumsum(np.linalg.norm(delta_new_cv, axis=1))
    s_accumulated = np.concatenate(([0], s_accumulated), axis=0)
    return new_cv, s_accumulated

def if_right_turning(entrance, exit):
    if entrance == 'w2' and exit == 's2':
        return True
    elif entrance == 's2' and exit == 'e2':
        return True
    elif entrance == 'e2' and exit == 'n2':
        return True
    elif entrance == 'n2' and exit == 'w2':
        return True
    else:
        return False

def if_left_turning(entrance, exit):
    if entrance == 'w1' and (exit == 'n1' or exit == 'n2'):
        return True
    elif entrance == 's1' and (exit == 'w2' or exit == 'w3'):
        return True
    elif entrance == 'e3' and (exit == 's2' or exit == 's3'):
        return True
    elif entrance == 'n3' and (exit == 'e1' or exit == 'e2'):
        return True
    else:
        return False

def if_going_straight(entrance, exit):
    if not if_right_turning(entrance, exit) and not if_left_turning(entrance, exit):
        return True
    else:
        return False

def intersection_ref_line(entrance, exit):  
    cv_init = None
    # north entrance
    if entrance == 'n2' and exit == 'w2':
        cv_init = np.array([[-7.5, 15], [-7.5, 13], [-9.5, 9.5], [-13, 7.5], [-15, 7.5]])
    if entrance == 'n2' and exit == 's2':
        cv_init = np.array([[-7.5, 15], [-7.5, 5], [-7.5, 0], [-7.5, -5], [-7.5, -15]])
    if entrance == 'n2' and exit == 's3':
        cv_init = np.array([[-7.5, 15], [-7, 10], [-5, 0], [-3, -10], [-2.5, -15]])
    if entrance == 'n3' and exit == 'e1':
        cv_init = np.array([[-2.5, 15], [-2.3, 13], [-1.8, 10], [2, 2], [10, -1.8], [13, -2.3], [15, -2.5]])
    if entrance == 'n3' and exit == 'e2':
        cv_init = np.array([[-2.5, 15], [-2.2, 10], [0, 0], [10, -6.8], [15, -7.5]])
    if entrance == 'n3' and exit == 's3':
        cv_init = np.array([[-2.5, 15], [-2.5, 5], [-2.5, 0], [-2.5, -5], [-2.5, -15]])

    # east entrance
    if entrance == 'e2' and exit == 'w3':
        cv_init = np.array([[15, 7.5], [10, 7], [0, 5], [-10, 3], [-15, 2.5]])
    if entrance == 'e2' and exit == 'w2':
        cv_init = np.array([[15, 7.5], [5, 7.5], [0, 7.5], [-5, 7.5], [-15, 7.5]])
    if entrance == 'e2' and exit == 'n2':
        cv_init = np.array([[15, 7.5], [13, 7.5], [9.5, 9.5], [7.5, 13], [7.5, 15]])
    if entrance == 'e3' and exit == 's3':
        cv_init = np.array([[15, 2.5], [12, 2.3], [10, 1.8], [2, -2], [-1.8, -10], [-2.3, -13], [-2.5, -15]])
    if entrance == 'e3' and exit == 's2':
        cv_init = np.array([[15, 2.5], [10, 2.2], [0, 0], [-6.8, -10], [-7.5, -15]])
    if entrance == 'e3' and exit == 'w3':
        cv_init = np.array([[15, 2.5], [5, 2.5], [0, 2.5], [-5, 2.5], [-15, 2.5]])

    # south entrance
    if entrance == 's2' and exit == 'n1':
        cv_init = np.array([[7.5, -15], [7, -10], [5, 0], [3, 10], [2.5, 15]])
    if entrance == 's2' and exit == 'n2':
        cv_init = np.array([[7.5, -15], [7.5, -5], [7.5, 0], [7.5, 5], [7.5, 15]])
    if entrance == 's2' and exit == 'e2':
        cv_init = np.array([[7.5, -15], [7.5, -13], [9.5, -9.5], [13, -7.5], [15, -7.5]])
    if entrance == 's1' and exit == 'w2':
        cv_init = np.array([[2.5, -15], [2.2, -10], [0, 0], [-10, 6.8], [-15, 7.5]])
    if entrance == 's1' and exit == 'w3':
        cv_init = np.array([[2.5, -15], [2.3, -13], [1.8, -10], [-2, -2], [-10, 1.8], [-13, 2.3], [-15, 2.5]])
    if entrance == 's1' and exit == 'n1':
        cv_init = np.array([[2.5, -15], [2.5, -5], [2.5, 0], [2.5, 5], [2.5, 15]])

    # west entrance
    if entrance == 'w1' and exit == 'n1':
        cv_init = np.array([[-15, -2.5], [-12, -2.3], [-10, -1.8], [-2, 2], [1.8, 10], [2.3, 13], [2.5, 15]])
    if entrance == 'w1' and exit == 'n2':
        cv_init = np.array([[-15, -2.5], [-10, -2.2], [0, 0], [6.8, 10], [7.5, 15]])
    if entrance == 'w1' and exit == 'e1':
        cv_init = np.array([[-15, -2.5], [-5, -2.5], [0, -2.5], [5, -2.5], [15, -2.5]])
    if entrance == 'w2' and exit == 's2':
        cv_init = np.array([[-15, -7.5], [-13, -7.5], [-9.5, -9.5], [-7.5, -13], [-7.5, -15]])
    if entrance == 'w2' and exit == 'e2':
        cv_init = np.array([[-15, -7.5], [-5, -7.5], [0, -7.5], [5, -7.5], [15, -7.5]])
    if entrance == 'w2' and exit == 'e1':
        cv_init = np.array([[-15, -7.5], [-10, -7], [0, -5], [10, -3], [15, -2.5]])

    assert cv_init is not None
    cv_smoothed, s_accumulated = smooth_ployline(cv_init)
    return cv_smoothed  # , s_accumulated

def record_ref_line_distance2exit(ref_line):
    """
    for calculate vehicle distance to exit which computation expensive
    store each point distance to exit
    only employ when T=0s or vehicle changes ref_line
    """
    cv = ref_line
    gap_list = np.zeros(len(cv))  
    for point in range(len(ref_line) - 1):
        gap = np.sqrt((cv[point, 0] - cv[point + 1, 0]) ** 2 +
                      (cv[point, 1] - cv[point + 1, 1]) ** 2)
        gap_list[point:] += gap
    ref_line_distance2exit = np.flipud(gap_list)
    return ref_line_distance2exit

def entrance_ref_line(entrance):
    """
    ref_line of entrance
    """
    ref_line = None
    # if entrance == 'n1':
    #     line = np.flipud(np.arange(15, 100, 0.01))
    #     ref_line = np.vstack((-12.5 * np.ones_like(line), line))
    if entrance == 'n2':
        line = np.flipud(np.arange(15, 100, 0.01))
        ref_line = np.vstack((-7.5 * np.ones_like(line), line))
    if entrance == 'n3':
        line = np.flipud(np.arange(15, 100, 0.01))
        ref_line = np.vstack((-2.5 * np.ones_like(line), line))
    # if entrance == 'e1':
    #     line = np.flipud(np.arange(15, 100, 0.01))
    #     ref_line = np.vstack((line, 12.5 * np.ones_like(line)))
    if entrance == 'e2':
        line = np.flipud(np.arange(15, 100, 0.01))
        ref_line = np.vstack((line, 7.5 * np.ones_like(line)))
    if entrance == 'e3':
        line = np.flipud(np.arange(15, 100, 0.01))
        ref_line = np.vstack((line, 2.5 * np.ones_like(line)))
    # if entrance == 's3':
    #     line = np.arange(-100, -15, 0.01)
    #     ref_line = np.vstack((12.5 * np.ones_like(line), line))
    if entrance == 's2':
        line = np.arange(-100, -15, 0.01)
        ref_line = np.vstack((7.5 * np.ones_like(line), line))
    if entrance == 's1':
        line = np.arange(-100, -15, 0.01)
        ref_line = np.vstack((2.5 * np.ones_like(line), line))
    # if entrance == 'w3':
    #     line = np.arange(-100, -15, 0.01)
    #     ref_line = np.vstack((line, -12.5 * np.ones_like(line)))
    if entrance == 'w2':
        line = np.arange(-100, -15, 0.01)
        ref_line = np.vstack((line, -7.5 * np.ones_like(line)))
    if entrance == 'w1':
        line = np.arange(-100, -15, 0.01)
        ref_line = np.vstack((line, -2.5 * np.ones_like(line)))
    return ref_line.T

def exit_ref_line(exit):
    """
    ref_line of exit
    """
    ref_line = None
    if exit == 'n1':
        line = np.arange(15, 100, 0.01)
        ref_line = np.vstack((2.5 * np.ones_like(line), line))
    if exit == 'n2':
        line = np.arange(15, 100, 0.01)
        ref_line = np.vstack((7.5 * np.ones_like(line), line))
    # if exit == 'n3':
    #     line = np.arange(15, 100, 0.01)
    #     ref_line = np.vstack((12.5 * np.ones_like(line), line))
    if exit == 'e1':
        line = np.arange(15, 100, 0.01)
        ref_line = np.vstack((line, -2.5 * np.ones_like(line)))
    if exit == 'e2':
        line = np.arange(15, 100, 0.01)
        ref_line = np.vstack((line, -7.5 * np.ones_like(line)))
    # if exit == 'e3':
    #     line = np.arange(15, 100, 0.01)
    #     ref_line = np.vstack((line, -12.5 * np.ones_like(line)))
    if exit == 's3':
        line = np.flipud(np.arange(-100, -15, 0.01))
        ref_line = np.vstack((-2.5 * np.ones_like(line), line))
    if exit == 's2':
        line = np.flipud(np.arange(-100, -15, 0.01))
        ref_line = np.vstack((-7.5 * np.ones_like(line), line))
    # if exit == 's1':
    #     line = np.flipud(np.arange(-100, -15, 0.01))
    #     ref_line = np.vstack((-12.5 * np.ones_like(line), line))
    if exit == 'w3':
        line = np.flipud(np.arange(-100, -15, 0.01))
        ref_line = np.vstack((line, 2.5 * np.ones_like(line)))
    if exit == 'w2':
        line = np.flipud(np.arange(-100, -15, 0.01))
        ref_line = np.vstack((line, 7.5 * np.ones_like(line)))
    # if exit == 'w1':
    #     line = np.flipud(np.arange(-100, -15, 0.01))
    #     ref_line = np.vstack((line, 12.5 * np.ones_like(line)))
    return ref_line.T

def concatenate_ref_lane(entrance, exit):
    ref_lane1 = entrance_ref_line(entrance)
    ref_lane2 = intersection_ref_line(entrance, exit)
    ref_lane3 = exit_ref_line(exit)
    ref_lane = np.vstack((ref_lane1, ref_lane2))
    ref_lane = np.vstack((ref_lane, ref_lane3))
    return ref_lane

def record_all_possible_ref_line():
    """
    instead of calculate ref_line at each time
    store all possible ref_line and its distance2exit(gap list)
    """
    possible_ref_line_list = []
    possible_ref_line_distance2exit_list = []
    total_length = []
    for entrance in POSSIBLE_ENTRANCE:
        entrance_possible_ref_line = []
        entrance_possible_ref_line_distance2exit = []
        entrance_possible_total_length = []
        for exit in ENTRANCE_EXIT_RELATION[entrance]:
            ref_line = concatenate_ref_lane(entrance, exit)
            entrance_possible_ref_line.append(ref_line)
            entrance_possible_ref_line_distance2exit.append(record_ref_line_distance2exit(ref_line))
            entrance_possible_total_length.append(max(record_ref_line_distance2exit(ref_line)))
        possible_ref_line_list.append(dict(zip(ENTRANCE_EXIT_RELATION[entrance], entrance_possible_ref_line)))
        possible_ref_line_distance2exit_list.append(dict(zip(ENTRANCE_EXIT_RELATION[entrance], entrance_possible_ref_line_distance2exit)))
        total_length.append(dict(zip(ENTRANCE_EXIT_RELATION[entrance], entrance_possible_total_length)))
    return dict(zip(POSSIBLE_ENTRANCE, possible_ref_line_list)), dict(zip(POSSIBLE_ENTRANCE, possible_ref_line_distance2exit_list)), dict(zip(POSSIBLE_ENTRANCE, total_length))

def update_pos_from_dis2des_to_Cartesian(state):
    entrance, exit = state['entrance'], state['exit']
    ref_line = ALL_REF_LINE[entrance][exit]
    gap_list = ALL_GAP_LIST[entrance][exit]
    index = np.argmin(abs(gap_list - state['dis2des']))
    x = ref_line[index, 0]
    y = ref_line[index, 1]
    return x, y

def default_exit_and_state(entrance, extra_direction):
    """
    default exit of each entrance
    """
    exit = None
    state = {}
    if entrance == 'n2':
        if extra_direction == 'straight':
            exit = 's2'
        else:
            exit = 'w2'
    if entrance == 'n3':
        if extra_direction == 'straight':
            exit = 's3'
        else:
            exit = 'e1'

    if entrance == 'e2':
        if extra_direction == 'straight':
            exit = 'w2'
        else:
            exit = 'n2'
    if entrance == 'e3':
        if extra_direction == 'straight':
            exit = 'w3'
        else:
            exit = 's3'

    if entrance == 's2':
        if extra_direction == 'straight':
            exit = 'n2'
        else:
            exit = 'e2'
    if entrance == 's1':
        if extra_direction == 'straight':
            exit = 'n1'
        else:
            exit = 'w3'

    if entrance == 'w2':
        if extra_direction == 'straight':
            exit = 'e2'
        else:
            exit = 's2'
    if entrance == 'w1':
        if extra_direction == 'straight':
            exit = 'e1'
        else:
            exit = 'n1'
    state['v'] = np.random.uniform(6, 9)
    state['heading'] = 0
    state['dis2des'] = ALL_GAP_LIST[entrance][exit][0] - np.random.randint(30, 60)
    state['entrance'] = entrance
    state['exit'] = exit
    state['x'], state['y'] = update_pos_from_dis2des_to_Cartesian(state)
    return exit, state

def calculate_heading(x1, y1, x2, y2):
    delta_x = x2 - x1
    delta_y = y2 - y1
    theta = math.atan2(delta_y, delta_x)
    heading = math.degrees(theta) % 360
    return heading


def kinematic_model(vehicle_info_old, acc):
    vehicle_info = copy.deepcopy(vehicle_info_old)
    vehicle_info['v'] = vehicle_info_old['v'] + acc * Dt
    vehicle_info['v'] = vehicle_info['v'] if vehicle_info['v'] < SPEED_LIMIT else SPEED_LIMIT
    vehicle_info['v'] = vehicle_info['v'] if vehicle_info['v'] > 0 else 0
    vehicle_info['dis2des'] = vehicle_info_old['dis2des'] - vehicle_info['v'] * Dt
    x, y = update_pos_from_dis2des_to_Cartesian(vehicle_info)
    vehicle_info['x'], vehicle_info['y'] = x, y
    vehicle_info['heading'] = calculate_heading(vehicle_info_old['x'], vehicle_info_old['y'], x, y)
    return vehicle_info

def get_dis2cp(vehicle1_info, vehicle2_info):
    dis2cp = vehicle1_info['dis2des'] - CONFLICT_RELATION[vehicle1_info['entrance']][vehicle1_info['exit']][str(vehicle2_info['entrance']) + str(vehicle2_info['exit'])]
    return dis2cp

def if_passed_conflict_point(ego_info, other_info):
    if str(ego_info['entrance']) + str(ego_info['exit']) in CONFLICT_RELATION[other_info['entrance']][other_info['exit']]:
        ego_dis2cp = get_dis2cp(ego_info, other_info)
        other_dis2cp = get_dis2cp(other_info, ego_info)
        if ego_dis2cp > 0 and other_dis2cp > 0:
            return False
        else:
            return True
    else:
        return True

def find_opponent(vehicle, state_list, front=False):  
    min_dis = 100000
    opponent = None
    for opp in state_list:
        if opp['id'] != vehicle['id'] and str(opp['entrance']) + str(opp['exit']) in CONFLICT_RELATION[vehicle['entrance']][vehicle['exit']]:
            veh_dis2cp = vehicle['dis2des'] - CONFLICT_RELATION[vehicle['entrance']][vehicle['exit']][str(opp['entrance']) + str(opp['exit'])]
            opp_dis2cp = opp['dis2des'] - CONFLICT_RELATION[opp['entrance']][opp['exit']][str(vehicle['entrance']) + str(vehicle['exit'])]
            if veh_dis2cp > 0 and opp_dis2cp > 0:
                if not front:  
                    if abs(veh_dis2cp - opp_dis2cp) < min_dis:
                        min_dis = abs(veh_dis2cp - opp_dis2cp)
                        opponent = opp
                else:  # for IDM
                    if 0 < veh_dis2cp - opp_dis2cp < min_dis:
                        min_dis = veh_dis2cp - opp_dis2cp
                        opponent = opp
    if opponent is not None:
        if opponent['v'] == 0 and front != True:
            opponent = find_opponent(vehicle, state_list, front=True)
    return opponent

def adjust_acc(vehicle, acc):
    if acc < 0 and REF_LINE_TOTAL_LENGTH[vehicle['entrance']][vehicle['exit']] - vehicle['dis2des'] < 70:  # 没过冲突点且要减速
        dis2stop_line = 70 - REF_LINE_TOTAL_LENGTH[vehicle['entrance']][vehicle['exit']] - vehicle['dis2des']
        acc2stop_line = vehicle['v'] ** 2 / (2 * dis2stop_line)
        if acc < acc2stop_line:
            acc = acc2stop_line
    return acc

def adjust_intention(vehicle, mcts_acc, game_acc, intention_remains):
    if vehicle['aggressive_intention'] != 10:
        if vehicle['v'] == 0 or game_acc < mcts_acc:  # only active for hdv
            if vehicle['aggressive_intention'] == -1:
                pass
            else:
                intention_remains = False
                vehicle['aggressive_intention'] = -1
        elif mcts_acc < game_acc:
            if vehicle['aggressive_intention'] == 1:
                pass
            else:
                intention_remains = False
                vehicle['aggressive_intention'] = 1
        else:
            pass
    return vehicle, intention_remains


def initialize_vehicle(veh_num):
    state_list = []
    filtered_entrance = copy.deepcopy(POSSIBLE_ENTRANCE)
    for veh in range(veh_num):
        add = False
        new_state = None
        while add is not True:
            new_entrance = random.choice(filtered_entrance)
            if not veh % 2 == 1:
                extra_direction = 'straight'
            else:
                extra_direction = 'turning'
            new_exit, new_state = default_exit_and_state(new_entrance, extra_direction)
            new_state['id'] = veh
            add = True
            new_state['type'] = 'cav' if random.randint(0, 10) < 11 else random.choice(['agg', 'con', 'nor'])
            for state in state_list:
                if str(new_state['entrance']) + str(new_state['exit']) in CONFLICT_RELATION[state['entrance']][state['exit']]:
                    if new_state['type'] == 'agg' and state['type'] == 'agg':
                        add = False
                # if state['entrance'] == new_state['entrance']:
                #     if new_state['dis2des'] > state['dis2des'] + 2 * VEH_L: 
                #         if new_state['dis2des'] - state['dis2des'] - 2 * VEH_L > (new_state['v'] ** 2 - state['v'] ** 2) / 2 * abs(MIN_ACCELERATION):
                #             add = False
                #     if state['dis2des'] > new_state['dis2des'] + 2 * VEH_L:
                #         if state['dis2des'] - new_state['dis2des'] - 2 * VEH_L > (state['v'] ** 2 - new_state['v'] ** 2) / 2 * abs(MIN_ACCELERATION):
                #             add = False
                #     else:
                #         add = False
            new_state['aggressive_intention'] = 0  
            new_state['heading'] = 0
            new_state['multiple_vehicle'] = 1  

        state_list.append(new_state)
        filtered_entrance = [entrance for entrance in filtered_entrance if entrance != new_entrance]
    return state_list

def conflict_point_occupied(vehicle, state_list):
    occupied = False
    for opp in state_list:
        if opp['id'] != vehicle['id'] and str(opp['entrance']) + str(opp['exit']) in CONFLICT_RELATION[vehicle['entrance']][vehicle['exit']]:
            veh_dis2cp = vehicle['dis2des'] - CONFLICT_RELATION[vehicle['entrance']][vehicle['exit']][str(opp['entrance']) + str(opp['exit'])]
            opp_dis2cp = opp['dis2des'] - CONFLICT_RELATION[opp['entrance']][opp['exit']][str(vehicle['entrance']) + str(vehicle['exit'])]
            if abs(opp_dis2cp) < 8 and veh_dis2cp > opp_dis2cp:
                occupied = True
                opp['aggressive_intention'] = 10
    return occupied, state_list

def open_excel(i, suffix, method):
    file_dir = './excel/' + strftime("%Y-%m-%d", gmtime()) + suffix + '-' + method + '/'
    file_name = file_dir + str(i) + '.xlsx'

    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
    workbook = openpyxl.Workbook()
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)

    # if 'Sheet' in workbook.sheetnames:
    #     del workbook['Sheet']
    return file_name, workbook

def write_data(workbook, state_list, t):
    column_names = ['t', 'x', 'y', 'v', 'theta', 'dis2des', 'entrance', 'exit', 'type', 'intention']
    for vehicle in state_list:
        sheet_name = str(vehicle['id'])
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(column_names)
        else:
            worksheet = workbook[sheet_name]
        state = [round(vehicle['x'], 2), round(vehicle['y'], 2), round(vehicle['v'], 2), round(vehicle['heading'], 2), round(vehicle['dis2des'], 2), vehicle['entrance'], vehicle['exit'], vehicle['type'], vehicle['aggressive_intention']]
        row_data = [t, round(vehicle['x'], 2), round(vehicle['y'], 2), round(vehicle['v'], 2), round(vehicle['heading'], 2), round(vehicle['dis2des'], 2), vehicle['entrance'], vehicle['exit'], vehicle['type'], vehicle['aggressive_intention']]
        worksheet.append(row_data)
        worksheet.cell(row=t + 2, column=1, value=t)
        for i, item in enumerate(state):
            worksheet.cell(row=t + 2, column=i + 2, value=item)
    return workbook

def cal_ttcp(vehicle, acc):
    veh_dis2cp, veh_v, veh_acc = vehicle['dis2des'], vehicle['v'], acc
    if veh_acc > 0:
        t_acc2max = (SPEED_LIMIT - veh_v) / veh_acc
        dis_acc2max = veh_v * t_acc2max + 0.5 * veh_acc * t_acc2max ** 2
        if dis_acc2max < veh_dis2cp:
            t_left = (veh_dis2cp - dis_acc2max) / SPEED_LIMIT
            ttcp = t_acc2max + t_left
        else:
            v = np.sqrt(veh_v ** 2 + 2 * veh_dis2cp * veh_acc)
            ttcp = (v - veh_v) / veh_acc
    elif veh_acc < 0:
        dis_acc2stop = veh_v ** 2 / (2 * abs(veh_acc))
        if dis_acc2stop < veh_dis2cp:
            ttcp = 10000
        else:
            v = np.sqrt(veh_v ** 2 + 2 * veh_dis2cp * veh_acc)
            ttcp = (v - veh_v) / veh_acc
    else:
        if veh_v == 0:
            ttcp = 10000
        else:
            ttcp = veh_dis2cp / veh_v
    return ttcp

def cal_instance_ttcp(vehicle, acc):
    veh_dis2cp, veh_v, veh_acc = vehicle['dis2des'], vehicle['v'], acc
    if veh_v == 0:
        return 10000
    ttcp = veh_dis2cp / veh_v
    return ttcp



'initialize map'
ALL_REF_LINE, ALL_GAP_LIST, REF_LINE_TOTAL_LENGTH = record_all_possible_ref_line() 
CONFLICT_RELATION = {'n2': {'w2': {'n2s2': 92.1900908028162, 'n2s3': 92.1900908028162, 'e2w2': 92.20009080284972, 's1w2': 92.20009080284972}, 's2': {'n2w2': 109.98999999990008, 'n2s3': 109.98999999990008, 'e2w2': 107.49750250083409, 'e2w3': 103.4061387129036, 'e3s2': 99.98999999999518, 'e3w3': 102.49583527842518, 's1w2': 106.08703234411477, 's1w3': 101.20540513504368, 'w2e1': 93.41280760253086, 'w2e2': 92.50250416805257, 'w2s2': 92.48999999999518, 'w1n1': 98.78459819939776, 'w1n2': 97.82427809269525, 'w1e1': 97.50417139046108}, 's3': {'n2w2': 110.44548312731945, 'n2s2': 110.44548312731945, 'n3s3': 84.98999999999518, 'e2w2': 107.89304730315321, 'e2w3': 104.04518589151061, 'e3s2': 96.0651736270741, 'e3s3': 99.98999999999518, 'e3w3': 102.79514988986395, 's1w2': 105.81252907056641, 's1w3': 100.21265729449242, 'w2e1': 94.25746224624378, 'w2e2': 92.56272521594435, 'w1n1': 100.24316290949524, 'w1n2': 98.53493745757231, 'w1e1': 97.66066298874941}}, 'n3': {'e1': {'n3e2': 105.91248683708801, 'n3s3': 105.91248683708801, 'e2w2': 105.56136937888002, 'e2w3': 102.78033402792454, 'e3s2': 97.80799431914734, 'e3s3': 95.6229977715223, 'e3w3': 99.64488293167916, 's2n1': 95.65961493394714, 's2n2': 92.86977862446395, 's1n1': 98.78569471469612, 'w2e1': 100.1262865919873, 'w1n1': 102.8078006399052, 'w1n2': 99.02989452006388, 'w1e1': 100.1262865919873}, 'e2': {'n3e1': 110.56444623292253, 'n3s3': 110.56444623292253, 'e2w2': 110.34917109494602, 'e2w3': 107.35444715559468, 'e3s2': 102.09650151542087, 'e3s3': 99.25785658032743, 'e3w3': 105.044561720145, 's2n1': 94.20640339205993, 's2n2': 92.79194988143217, 's2e2': 92.59317962558931, 's1n1': 98.65972688258648, 'w2e1': 96.66792576418933, 'w2e2': 100.09317962558931, 'w1n1': 105.86197541301765, 'w1n2': 102.53370847322104, 'w1e1': 99.01653786837544}, 's3': {'n2s3': 84.98999999999518, 'n3e1': 107.48999999990008, 'n3e2': 107.48999999990008, 'e2w2': 107.49750250083409, 'e2w3': 104.43648216071983, 'e3s2': 97.76425808602635, 'e3s3': 99.98999999999518, 'e3w3': 102.49583527842518, 's1w2': 102.75592197399044, 's1w3': 98.10437145715015, 'w2e1': 94.43314771590215, 'w2e2': 92.50250416805257, 'w1n1': 101.90563854618092, 'w1n2': 99.01467489162857, 'w1e1': 97.50417139046108}}, 'e2': {'n2': {'e2w2': 92.18169159380608, 'e2w3': 92.18169159380608, 's2n2': 92.19169159386078, 'w1n2': 92.19169159386078}, 'w2': {'n2w2': 92.48999999999518, 'n2s2': 92.50250416805257, 'n2s3': 93.41280760253086, 'n3e1': 98.79460153384258, 'n3e2': 97.82427809269525, 'n3s3': 97.50417139046108, 'e2n2': 109.98999999990008, 'e2w3': 109.98999999990008, 's2n1': 103.4061387129036, 's2n2': 107.49750250083409, 's1w2': 99.98999999999518, 's1n1': 102.49583527842518, 'w1n1': 101.20540513504368, 'w1n2': 106.08703234411477}, 'w3': {'n2s2': 92.56272521594435, 'n2s3': 94.25746224624378, 'n3e1': 100.24316290949524, 'n3e2': 98.53493745757231, 'n3s3': 97.66066298874941, 'e2n2': 110.44548312731945, 'e2w2': 110.44548312731945, 'e3w3': 84.98999999999518, 's2n1': 104.04518589151061, 's2n2': 107.89304730315321, 's1w2': 96.0651736270741, 's1w3': 99.98999999999518, 's1n1': 102.79514988986395, 'w1n1': 100.21265729449242, 'w1n2': 105.81252907056641}}, 'e3': {'s2': {'n2s2': 100.09317962558931, 'n2s3': 96.66792576418933, 'n3e1': 105.85104894583496, 'n3e2': 102.53370847322104, 'n3s3': 99.01653786837544, 'e3s3': 110.55648181606591, 'e3w3': 110.55648181606591, 's2n1': 107.35444715559468, 's2n2': 110.34917109494602, 's1w2': 102.09650151542087, 's1w3': 99.25785658032743, 's1n1': 105.044561720145, 'w2e1': 94.20640339205993, 'w2e2': 92.79194988143217, 'w2s2': 92.59317962558931, 'w1e1': 98.65972688258648}, 's3': {'n2s3': 100.10309672481955, 'n3e1': 102.76635512219673, 'n3e2': 99.00843389356224, 'n3s3': 100.10309672481955, 'e3s2': 105.89496288611605, 'e3w3': 105.89496288611605, 's2n1': 102.7388828133214, 's2n2': 105.5183496613374, 's1w2': 97.77209183981026, 's1w3': 95.58534511487673, 's1n1': 99.6074502201253, 'w2e1': 95.62214057020425, 'w2e2': 92.82125214701615, 'w1e1': 98.7632620916063}, 'w3': {'n2s2': 92.50250416805257, 'n2s3': 94.43314771590215, 'n3e1': 101.8956352117361, 'n3e2': 99.01467489162857, 'n3s3': 97.50417139046108, 'e2w3': 84.98999999999518, 'e3s2': 107.48999999990008, 'e3s3': 107.48999999990008, 's2n1': 104.43648216071983, 's2n2': 107.49750250083409, 's1w2': 97.76425808602635, 's1w3': 99.98999999999518, 's1n1': 102.49583527842518, 'w1n1': 98.09436812270533, 'w1n2': 102.75592197399044}}, 's2': {'n1': {'n3e1': 100.22265729449724, 'n3e2': 105.82252907057124, 'e2w2': 92.57272521594918, 'e2w3': 94.2674622462486, 'e3s2': 98.54493745757713, 'e3s3': 100.25316290950006, 'e3w3': 97.67066298875423, 's2n2': 110.44548312735958, 's2e2': 110.44548312735958, 's1n1': 85.0, 'w2e1': 104.05518589151544, 'w2e2': 107.90304730315803, 'w1n1': 100.0, 'w1n2': 96.07517362707893, 'w1e1': 102.80514988986877}, 'n2': {'n3e1': 101.2154051350485, 'n3e2': 106.0970323441196, 'e2n2': 92.5, 'e2w2': 92.51250416805739, 'e2w3': 93.42280760253568, 'e3s2': 97.83427809270007, 'e3s3': 98.79459819940259, 'e3w3': 97.5141713904659, 's2n1': 109.98999999993981, 's2e2': 109.98999999993981, 'w2e1': 103.41613871290842, 'w2e2': 107.50750250083891, 'w1n2': 100.0, 'w1e1': 102.50583527843}, 'e2': {'n3e2': 92.20009080289802, 's2n1': 92.19009080284332, 's2n2': 92.19009080284332, 'w2e2': 92.20009080289802}}, 's1': {'w2': {'n2w2': 92.58774853573118, 'n2s2': 92.78651879157404, 'n2s3': 94.20097230220179, 'n3s3': 98.65429579272835, 'e2w2': 100.08774853573118, 'e2w3': 96.6624946743312, 'e3s2': 102.5282773833629, 'e3s3': 105.85654432315951, 'e3w3': 99.0111067785173, 's1w3': 110.55105072620773, 's1n1': 110.55105072620773, 'w2e1': 107.34901606573655, 'w2e2': 110.34374000508788, 'w1n1': 99.25242549046929, 'w1n2': 102.09107042556273, 'w1e1': 105.03913063028686}, 'w3': {'n2s2': 92.86372113183914, 'n2s3': 95.65355744132233, 'n3s3': 98.77963722207132, 'e2w3': 100.12022909936249, 'e3s2': 99.02383702743907, 'e3s3': 102.80174314728039, 'e3w3': 100.12022909936249, 's1w2': 105.90037185179978, 's1n1': 105.90037185179978, 'w2e1': 102.77427653529973, 'w2e2': 105.55531188625521, 'w1n1': 95.61694027889749, 'w1n2': 97.80193682652254, 'w1e1': 99.63882543905436}, 'n1': {'n3e1': 98.11437145715497, 'n3e2': 102.76592197399526, 'e2w2': 92.51250416805739, 'e2w3': 94.44314771590697, 'e3s2': 99.0246748916334, 'e3s3': 101.91563854618575, 'e3w3': 97.5141713904659, 's2n1': 85.0, 's1w2': 107.48999999993981, 's1w3': 107.48999999993981, 'w2e1': 104.44648216072466, 'w2e2': 107.50750250083891, 'w1n1': 100.0, 'w1n2': 97.77425808603117, 'w1e1': 102.50583527843}}, 'w2': {'e1': {'n2s2': 107.90304730315803, 'n2s3': 104.05518589151544, 'n3e1': 100.0, 'n3e2': 96.07517362707893, 'n3s3': 102.80514988986877, 'e3s2': 105.82252907057124, 'e3s3': 100.22265729449724, 's2n1': 94.2674622462486, 's2n2': 92.57272521594918, 's1w2': 98.54493745757713, 's1w3': 100.25316290950006, 's1n1': 97.67066298875423, 'w2e2': 110.44548312735958, 'w2s2': 110.44548312735958, 'w1e1': 85.0}, 'e2': {'n2s2': 107.50750250083891, 'n2s3': 103.41613871290842, 'n3e2': 100.0, 'n3s3': 102.50583527843, 'e3s2': 106.0970323441196, 'e3s3': 101.2154051350485, 's2n1': 93.42280760253568, 's2n2': 92.51250416805739, 's2e2': 92.5, 's1w2': 97.83427809270007, 's1w3': 98.8046015338474, 's1n1': 97.5141713904659, 'w2e1': 109.98999999993981, 'w2s2': 109.98999999993981}, 's2': {'n2s2': 92.20849001188695, 'e3s2': 92.20849001188695, 'w2e1': 92.19849001185344, 'w2e2': 92.19849001185344}}, 'w1': {'n1': {'n2s2': 105.51110240417097, 'n2s3': 102.73163555615497, 'n3e1': 95.5780978577103, 'n3e2': 97.76484458264383, 'n3s3': 99.60020296295887, 'e2w2': 92.81400488984973, 'e2w3': 95.61489331303783, 'e3w3': 98.75601483443987, 's2n1': 100.09584946765312, 's1w2': 99.00118663639581, 's1w3': 102.7591078650303, 's1n1': 100.09584946765312, 'w1n2': 105.89373257569889, 'w1e1': 105.89373257569889}, 'n2': {'n2s2': 110.34374000508788, 'n2s3': 107.34901606573655, 'n3e1': 99.25242549046929, 'n3e2': 102.09107042556273, 'n3s3': 105.03913063028686, 'e2n2': 92.58774853573118, 'e2w2': 92.78651879157404, 'e2w3': 94.20097230220179, 'e3w3': 98.65429579272835, 's2n1': 96.6624946743312, 's2n2': 100.08774853573118, 's1w2': 102.5282773833629, 's1w3': 105.84561785597683, 's1n1': 99.0111067785173, 'w1n1': 110.55901514306441, 'w1e1': 110.55901514306441}, 'e1': {'n2s2': 107.50750250083891, 'n2s3': 104.44648216072466, 'n3e1': 100.0, 'n3e2': 97.77425808603117, 'n3s3': 102.50583527843, 'e3s2': 102.76592197399526, 'e3s3': 98.10436812271016, 's2n1': 94.44314771590697, 's2n2': 92.51250416805739, 's1w2': 99.0246748916334, 's1w3': 101.90563521174093, 's1n1': 97.5141713904659, 'w2e1': 85.0, 'w1n1': 107.48999999993981, 'w1n2': 107.48999999993981}}}
CONFLICT_RELATION_STATE = {'n2': {'w2': {'n2s2': (-8.76384618588846, 10.30610693839657), 'n2s3': (-8.76384618588846, 10.30610693839657), 'e2w2': (-8.756395164701242, 10.316049319054008), 's1w2': (-8.756395164701242, 10.316049319054008)}, 's2': {'n2w2': (-7.499999999999998, 9.988329443147714), 'n2s3': (-7.499999999999998, 9.988329443147714), 'e2w2': (-7.5, 7.497499166388798), 'e2w3': (-7.5, 3.4061353784594877), 'e3s2': (-7.5, -0.015005001667222073), 'e3w3': (-7.500000000000001, 2.4958319439813277), 's1w2': (-7.499999999999999, 6.08702900966989), 's1w3': (-7.499999999999998, 1.2054018006001992), 'w2e1': (-7.5, -6.587195731910638), 'w2e2': (-7.5, -7.497499166388796), 'w2s2': (-7.499999999999999, -7.50750250083361), 'w1n1': (-7.500000000000001, -1.2154051350450157), 'w1n2': (-7.499999999999999, -2.175725241747248), 'w1e1': (-7.5, -2.495831943981327)}, 's3': {'n2w2': (-7.002878461283799, 10.019752095103529), 'n2s2': (-7.002878461283799, 10.019752095103529), 'n3s3': (-2.5, -15.00999999995652), 'e2w2': (-6.588874524618683, 7.497993643546099), 'e2w3': (-5.8327595544662385, 3.7253200853767767), 'e3s2': (-4.094692424756606, -4.063059891311384), 'e3s3': (-4.947021072841093, -0.23301352812398823), 'e3w3': (-5.565072915113688, 2.5042797807949295), 's1w2': (-6.1955429996019165, 5.455023860975121), 's1w3': (-4.996618150848345, -0.014873171676174968), 'w2e1': (-3.7284112707325, -5.833251521550004), 'w2e2': (-3.4111254753813163, -7.497993643546108), 'w1n1': (-5.003381849151657, 0.014873171676168973), 'w1n2': (-4.625820355350548, -1.651102456631883), 'w1e1': (-4.434927084886315, -2.5042797807949344)}}, 'n3': {'e1': {'n3e2': (-1.3265168367278375, 7.831771102535564), 'n3s3': (-1.3265168367278375, 7.831771102535564), 'e2w2': (-1.2022311995118957, 7.499211853982482), 'e2w3': (0.01174524168369584, 5.00110710144212), 'e3s2': (3.235647332708594, 1.2488003877627587), 'e3s3': (5.024684179909494, -0.002342117016498148), 'e3w3': (1.8928646207888682, 2.5003036883979757), 's2n1': (4.9932545021648895, 0.01644983060813953), 's2n2': (7.49921185398248, -1.2022311995118964), 's1n1': (2.500303688397977, 1.8928646207888664), 'w2e1': (1.5705752190551652, 2.8553914340394555), 'w1n1': (-0.002342117016496781, 5.024684179909494), 'w1n2': (2.323814099613893, 2.061637233102535), 'w1e1': (1.5705752190551652, 2.8553914340394555)}, 'e2': {'n3e1': (-2.201211133735817, 7.7099040943674915), 'n3s3': (-2.201211133735817, 7.7099040943674915), 'e2w2': (-2.176359792364172, 7.50305574699929), 'e2w3': (-1.653485824193486, 4.630096907254525), 'e3s2': (0.2994772454442032, -0.07385788758142253), 'e3s3': (2.0656784922423608, -2.323739128393538), 'e3w3': (-0.9852136256671373, 2.49731250504746), 's2n1': (6.199487368923628, -5.454175611859762), 's2n2': (7.499922728263725, -6.084901874469155), 's2e2': (7.6845127427308375, -6.1638014330224635), 's1n1': (2.4987131196229178, -2.759878397602499), 'w2e1': (4.0633818724914805, -4.092200208173672), 'w2e2': (1.4910021659346708, -1.6866079150160989), 'w1n1': (-1.2520216686583794, 3.2420007451734016), 'w1n2': (0.07403523969618941, 0.2944483676375764), 'w1e1': (2.238112111069683, -2.501550489944223)}, 's3': {'n2s3': (-2.5, -15.00999999995652), 'n3e1': (-2.4999999999999982, 7.487495831943981), 'n3e2': (-2.4999999999999982, 7.487495831943981), 'e2w2': (-2.4999999999999996, 7.497499166388798), 'e2w3': (-2.499999999999999, 4.436478826275426), 'e3s2': (-2.4999999999999996, -2.235745248416137), 'e3s3': (-2.5, -0.015005001667222073), 'e3w3': (-2.5, 2.4958319439813277), 's1w2': (-2.499999999999999, 2.7559186395465156), 's1w3': (-2.5, -1.8956318772924305), 'w2e1': (-2.5, -5.566855618539513), 'w2e2': (-2.5000000000000004, -7.497499166388796), 'w1n1': (-2.4999999999999996, 1.905635211737247), 'w1n2': (-2.5, -0.9853284428142701), 'w1e1': (-2.5, -2.495831943981327)}}, 'e2': {'n2': {'e2w2': (10.299484972978993, 8.768822913920046), 'e2w3': (10.299484972978993, 8.768822913920046), 's2n2': (10.309419808783607, 8.761360634872663), 'w1n2': (10.309419808783607, 8.761360634872663)}, 'w2': {'n2w2': (-7.50750250083361, 7.499999999999999), 'n2s2': (-7.497499166388796, 7.5), 'n2s3': (-6.587195731910638, 7.5), 'n3e1': (-1.2054018006001996, 7.499999999999999), 'n3e2': (-2.175725241747248, 7.499999999999999), 'n3s3': (-2.495831943981327, 7.5), 'e2n2': (9.988329443147714, 7.499999999999998), 'e2w3': (9.988329443147714, 7.499999999999998), 's2n1': (3.4061353784594877, 7.5), 's2n2': (7.497499166388798, 7.5), 's1w2': (-0.015005001667222073, 7.5), 's1n1': (2.4958319439813277, 7.500000000000001), 'w1n1': (1.2054018006001992, 7.499999999999998), 'w1n2': (6.08702900966989, 7.499999999999999)}, 'w3': {'n2s2': (-7.497993643546108, 3.4111254753813163), 'n2s3': (-5.833251521550004, 3.7284112707325), 'n3e1': (0.014873171676168973, 5.003381849151657), 'n3e2': (-1.651102456631883, 4.625820355350548), 'n3s3': (-2.5042797807949344, 4.434927084886315), 'e2n2': (10.019752095103529, 7.002878461283799), 'e2w2': (10.019752095103529, 7.002878461283799), 'e3w3': (-15.00999999995652, 2.5), 's2n1': (3.7253200853767767, 5.8327595544662385), 's2n2': (7.497993643546099, 6.588874524618683), 's1w2': (-4.063059891311384, 4.094692424756606), 's1w3': (-0.23301352812398823, 4.947021072841093), 's1n1': (2.5042797807949295, 5.565072915113688), 'w1n1': (-0.014873171676174968, 4.996618150848345), 'w1n2': (5.455023860975121, 6.1955429996019165)}}, 'e3': {'s2': {'n2s2': (-1.6866079150160989, -1.4910021659346708), 'n2s3': (-4.092200208173672, -4.0633818724914805), 'n3e1': (3.2319690964096903, 1.2486843549020696), 'n3e2': (0.2944483676375764, -0.07403523969618941), 'n3s3': (-2.501550489944223, -2.238112111069683), 'e3s3': (7.7099040943674915, 2.201211133735817), 'e3w3': (7.7099040943674915, 2.201211133735817), 's2n1': (4.630096907254525, 1.653485824193486), 's2n2': (7.50305574699929, 2.176359792364172), 's1w2': (-0.07385788758142253, -0.2994772454442032), 's1w3': (-2.323739128393538, -2.0656784922423608), 's1n1': (2.49731250504746, 0.9852136256671373), 'w2e1': (-5.454175611859762, -6.199487368923628), 'w2e2': (-6.084901874469155, -7.499922728263725), 'w2s2': (-6.1638014330224635, -7.6845127427308375), 'w1e1': (-2.759878397602499, -2.4987131196229178)}, 's3': {'n2s3': (2.867796793286925, -1.5685492116673074), 'n3e1': (5.0221773483456955, 0.0006430442386728963), 'n3e2': (2.0705826567763683, -2.324541550183417), 'n3s3': (2.867796793286925, -1.5685492116673074), 'e3s2': (7.850599991720309, 1.3496043988826767), 'e3w3': (7.850599991720309, 1.3496043988826767), 's2n1': (4.998546881567843, -0.013606730716534399), 's2n2': (7.500326254422595, 1.216870315845671), 's1w2': (1.2486285645637287, -3.2450625347241253), 's1w3': (-0.0013533097611743705, -5.027139798596033), 's1n1': (2.4979186031305023, -1.9046343225240465), 'w2e1': (0.01750070678960703, -4.995731886399233), 'w2e2': (-1.2044975729414038, -7.498864348790106), 'w1e1': (1.9010920603710877, -2.5014771323536324)}, 'w3': {'n2s2': (-7.497499166388796, 2.5000000000000004), 'n2s3': (-5.566855618539513, 2.5), 'n3e1': (1.8956318772924305, 2.4999999999999987), 'n3e2': (-0.9853284428142701, 2.5), 'n3s3': (-2.495831943981327, 2.5), 'e2w3': (-15.00999999995652, 2.5), 'e3s2': (7.487495831943981, 2.4999999999999982), 'e3s3': (7.487495831943981, 2.4999999999999982), 's2n1': (4.436478826275426, 2.499999999999999), 's2n2': (7.497499166388798, 2.4999999999999996), 's1w2': (-2.235745248416137, 2.4999999999999996), 's1w3': (-0.015005001667222073, 2.5), 's1n1': (2.4958319439813277, 2.5), 'w1n1': (-1.9056352117372453, 2.4999999999999996), 'w1n2': (2.7559186395465156, 2.499999999999999)}}, 's2': {'n1': {'n3e1': (4.996618150848345, 0.014873171676174968), 'n3e2': (6.1955429996019165, -5.455023860975121), 'e2w2': (3.4111254753813163, 7.497993643546108), 'e2w3': (3.7284112707325, 5.833251521550004), 'e3s2': (4.625820355350548, 1.651102456631883), 'e3s3': (5.003381849151657, -0.014873171676168973), 'e3w3': (4.434927084886315, 2.5042797807949344), 's2n2': (7.00141697470169, -10.009717975558141), 's2e2': (7.00141697470169, -10.009717975558141), 's1n1': (2.5, 15.0), 'w2e1': (5.8327595544662385, -3.7253200853767767), 'w2e2': (6.588874524618683, -7.497993643546099), 'w1n1': (4.947021072841093, 0.23301352812398823), 'w1n2': (4.094692424756606, 4.063059891311384), 'w1e1': (5.565072915113688, -2.5042797807949295)}, 'n2': {'n3e1': (7.499999999999998, -1.2054018006001992), 'n3e2': (7.499999999999999, -6.08702900966989), 'e2n2': (7.499999999999999, 7.50750250083361), 'e2w2': (7.5, 7.497499166388796), 'e2w3': (7.5, 6.587195731910638), 'e3s2': (7.499999999999999, 2.175725241747248), 'e3s3': (7.500000000000001, 1.2154051350450157), 'e3w3': (7.5, 2.495831943981327), 's2n1': (7.5, -9.978326108702902), 's2e2': (7.5, -9.978326108702902), 'w2e1': (7.5, -3.4061353784594877), 'w2e2': (7.5, -7.497499166388798), 'w1n2': (7.5, 0.015005001667222073), 'w1e1': (7.500000000000001, -2.4958319439813277)}, 'e2': {'n3e2': (8.761360634872663, -10.309419808783607), 's2n1': (8.768822913920046, -10.299484972978993), 's2n2': (8.768822913920046, -10.299484972978993), 'w2e2': (8.761360634872663, -10.309419808783607)}}, 's1': {'w2': {'n2w2': (-7.6845127427308375, 6.1638014330224635), 'n2s2': (-7.499922728263725, 6.084901874469155), 'n2s3': (-6.199487368923628, 5.454175611859762), 'n3s3': (-2.4987131196229178, 2.759878397602499), 'e2w2': (-1.4910021659346708, 1.6866079150160989), 'e2w3': (-4.0633818724914805, 4.092200208173672), 'e3s2': (-0.07403523969618941, -0.2944483676375764), 'e3s3': (1.2520216686583794, -3.2420007451734016), 'e3w3': (-2.238112111069683, 2.501550489944223), 's1w3': (2.201211133735817, -7.7099040943674915), 's1n1': (2.201211133735817, -7.7099040943674915), 'w2e1': (1.653485824193486, -4.630096907254525), 'w2e2': (2.176359792364172, -7.50305574699929), 'w1n1': (-2.0656784922423608, 2.323739128393538), 'w1n2': (-0.2994772454442032, 0.07385788758142253), 'w1e1': (0.9852136256671373, -2.49731250504746)}, 'w3': {'n2s2': (-7.49921185398248, 1.2022311995118964), 'n2s3': (-4.9932545021648895, -0.01644983060813953), 'n3s3': (-2.500303688397977, -1.8928646207888664), 'e2w3': (-1.5705752190551652, -2.8553914340394555), 'e3s2': (-2.323814099613893, -2.061637233102535), 'e3s3': (0.002342117016496781, -5.024684179909494), 'e3w3': (-1.5705752190551652, -2.8553914340394555), 's1w2': (1.323330332163208, -7.8229755342139615), 's1n1': (1.323330332163208, -7.8229755342139615), 'w2e1': (-0.01174524168369584, -5.00110710144212), 'w2e2': (1.2022311995118957, -7.499211853982482), 'w1n1': (-5.024684179909494, 0.002342117016498148), 'w1n2': (-3.235647332708594, -1.2488003877627587), 'w1e1': (-1.8928646207888682, -2.5003036883979757)}, 'n1': {'n3e1': (2.5, 1.8956318772924305), 'n3e2': (2.499999999999999, -2.7559186395465156), 'e2w2': (2.5000000000000004, 7.497499166388796), 'e2w3': (2.5, 5.566855618539513), 'e3s2': (2.5, 0.9853284428142701), 'e3s3': (2.4999999999999996, -1.905635211737247), 'e3w3': (2.5, 2.495831943981327), 's2n1': (2.5, 15.0), 's1w2': (2.499999999999999, -7.477492497499165), 's1w3': (2.499999999999999, -7.477492497499165), 'w2e1': (2.499999999999999, -4.436478826275426), 'w2e2': (2.4999999999999996, -7.497499166388798), 'w1n1': (2.5, 0.015005001667222073), 'w1n2': (2.4999999999999996, 2.235745248416137), 'w1e1': (2.5, -2.4958319439813277)}}, 'w2': {'e1': {'n2s2': (-7.497993643546099, -6.588874524618683), 'n2s3': (-3.7253200853767767, -5.8327595544662385), 'n3e1': (0.23301352812398823, -4.947021072841093), 'n3e2': (4.063059891311384, -4.094692424756606), 'n3s3': (-2.5042797807949295, -5.565072915113688), 'e3s2': (-5.455023860975121, -6.1955429996019165), 'e3s3': (0.014873171676174968, -4.996618150848345), 's2n1': (5.833251521550004, -3.7284112707325), 's2n2': (7.497993643546108, -3.4111254753813163), 's1w2': (1.651102456631883, -4.625820355350548), 's1w3': (-0.014873171676168973, -5.003381849151657), 's1n1': (2.5042797807949344, -4.434927084886315), 'w2e2': (-10.009717975558141, -7.00141697470169), 'w2s2': (-10.009717975558141, -7.00141697470169), 'w1e1': (15.0, -2.5)}, 'e2': {'n2s2': (-7.497499166388798, -7.5), 'n2s3': (-3.4061353784594877, -7.5), 'n3e2': (0.015005001667222073, -7.5), 'n3s3': (-2.4958319439813277, -7.500000000000001), 'e3s2': (-6.08702900966989, -7.499999999999999), 'e3s3': (-1.2054018006001992, -7.499999999999998), 's2n1': (6.587195731910638, -7.5), 's2n2': (7.497499166388796, -7.5), 's2e2': (7.50750250083361, -7.499999999999999), 's1w2': (2.175725241747248, -7.499999999999999), 's1w3': (1.2054018006001996, -7.499999999999999), 's1n1': (2.495831943981327, -7.5), 'w2e1': (-9.978326108702902, -7.5), 'w2s2': (-9.978326108702902, -7.5)}, 's2': {'n2s2': (-10.316049319054008, -8.756395164701242), 'e3s2': (-10.316049319054008, -8.756395164701242), 'w2e1': (-10.30610693839657, -8.76384618588846), 'w2e2': (-10.30610693839657, -8.76384618588846)}}, 'w1': {'n1': {'n2s2': (-7.500326254422595, -1.216870315845671), 'n2s3': (-4.998546881567843, 0.013606730716534399), 'n3e1': (0.0013533097611743705, 5.027139798596033), 'n3e2': (-1.2486285645637287, 3.2450625347241253), 'n3s3': (-2.4979186031305023, 1.9046343225240465), 'e2w2': (1.2044975729414038, 7.498864348790106), 'e2w3': (-0.01750070678960703, 4.995731886399233), 'e3w3': (-1.9010920603710877, 2.5014771323536324), 's2n1': (-2.867796793286925, 1.5685492116673074), 's1w2': (-2.0705826567763683, 2.324541550183417), 's1w3': (-5.0221773483456955, -0.0006430442386728963), 's1n1': (-2.867796793286925, 1.5685492116673074), 'w1n2': (-7.859404025179848, -1.3528287410422744), 'w1e1': (-7.859404025179848, -1.3528287410422744)}, 'n2': {'n2s2': (-7.50305574699929, -2.176359792364172), 'n2s3': (-4.630096907254525, -1.653485824193486), 'n3e1': (2.323739128393538, 2.0656784922423608), 'n3e2': (0.07385788758142253, 0.2994772454442032), 'n3s3': (-2.49731250504746, -0.9852136256671373), 'e2n2': (6.1638014330224635, 7.6845127427308375), 'e2w2': (6.084901874469155, 7.499922728263725), 'e2w3': (5.454175611859762, 6.199487368923628), 'e3w3': (2.759878397602499, 2.4987131196229178), 's2n1': (4.092200208173672, 4.0633818724914805), 's2n2': (1.6866079150160989, 1.4910021659346708), 's1w2': (-0.2944483676375764, 0.07403523969618941), 's1w3': (-3.2319690964096903, -1.2486843549020696), 's1n1': (2.501550489944223, 2.238112111069683), 'w1n1': (-7.7099040943674915, -2.201211133735817), 'w1e1': (-7.7099040943674915, -2.201211133735817)}, 'e1': {'n2s2': (-7.497499166388798, -2.4999999999999996), 'n2s3': (-4.436478826275426, -2.499999999999999), 'n3e1': (0.015005001667222073, -2.5), 'n3e2': (2.235745248416137, -2.4999999999999996), 'n3s3': (-2.4958319439813277, -2.5), 'e3s2': (-2.7559186395465156, -2.499999999999999), 'e3s3': (1.9056352117372453, -2.4999999999999996), 's2n1': (5.566855618539513, -2.5), 's2n2': (7.497499166388796, -2.5000000000000004), 's1w2': (0.9853284428142701, -2.5), 's1w3': (-1.8956318772924305, -2.4999999999999987), 's1n1': (2.495831943981327, -2.5), 'w2e1': (15.0, -2.5), 'w1n1': (-7.477492497499165, -2.499999999999999), 'w1n2': (-7.477492497499165, -2.499999999999999)}}}
